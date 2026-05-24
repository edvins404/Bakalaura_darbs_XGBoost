import io
import time
import argparse
import joblib
import numpy as np
import pandas as pd
from pathlib import Path
from xgboost import XGBClassifier
from xgboost.callback import TrainingCallback
from tqdm import tqdm
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import (
    classification_report,
    confusion_matrix,
    accuracy_score,
    f1_score,
    roc_auc_score,
    precision_score,
)
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage

RESULTS_FILE = Path("rezultati_kopsavilkums.xlsx")
SUMMARY_SHEET_NAME = "Kopsavilkums"

XGBOOST_PARAMS = {
    "n_estimators": 300,
    "max_depth": 6,
    "learning_rate": 0.1,
    "subsample": 0.8,
    "colsample_bytree": 0.8,
    "eval_metric": "mlogloss",
    "random_state": 42,
    "n_jobs": -1,
    "tree_method": "hist",  
    "device": "cuda",
}

MODEL_NAME_STRIP_PREFIX = "merged_no_nan_clean_headers_no_duplicates_train"

SUMMARY_HEADERS = [
    "Dataset",
    "Accuracy",
    "Precision Macro",
    "F1 Macro",
    "F1 Weighted",
    "AUC-ROC",
    "Prediction Time (s)",
    "Records",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
ROW_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# Atvasina enkodiera faila ceļu no modeļa faila ceļa
def resolve_encoder_path(model_path: str) -> Path:
    p = Path(model_path)
    return p.parent / f"{p.stem}_encoder.pkl"


# Ielādē CSV failu un sadala datos pazīmēs un mērķa mainīgajā
def load_dataset(csv_path: str, target_column: str) -> tuple[pd.DataFrame, pd.Series]:
    dataframe = pd.read_csv(csv_path, low_memory=False)
    features = dataframe.drop(columns=[target_column])
    labels = dataframe[target_column]
    return features, labels


# Enkodē klašu etiķetes skaitliskās vērtībās un saglabā enkoderi atkārtotai izmantošanai
def encode_labels(labels: pd.Series, encoder_path: Path) -> tuple[np.ndarray, LabelEncoder]:
    encoder = LabelEncoder()
    encoded_labels = encoder.fit_transform(labels)
    joblib.dump(encoder, encoder_path)
    return encoded_labels, encoder


# Aprēķina logaritmiski nogludinātos klases svarus, lai mazinātu nelīdzsvarotu klašu ietekmi
def compute_sample_weights(encoded_labels: np.ndarray) -> np.ndarray:
    class_counts = np.bincount(encoded_labels)
    total_samples = len(encoded_labels)
    smoothed_weights = np.log(1 + (total_samples / (class_counts))) 
    return smoothed_weights[encoded_labels]


# Savieno XGBoost apmācības iterācijas ar tqdm progresa joslu
class TqdmProgressCallback(TrainingCallback):
    def __init__(self, total_trees: int):
        self._progress_bar = tqdm(total=total_trees, desc="Apmaciba", unit="koks")

    def after_iteration(self, model, epoch, evals_log) -> bool:
        self._progress_bar.update(1)
        return False

    def after_training(self, model):
        self._progress_bar.close()
        return model


# Ģenerē konfūzijas matricas attēlu atmiņā, lai iegultu Excel failā bez starpfailiem
def build_confusion_matrix_image(
    true_labels: np.ndarray,
    predicted_labels: np.ndarray,
    class_names: list[str],
) -> io.BytesIO:
    confusion = confusion_matrix(true_labels, predicted_labels)

    plt.figure(figsize=(14, 10))
    sns.heatmap(
        confusion,
        annot=True,
        fmt="d",
        cmap="Blues",
        xticklabels=class_names,
        yticklabels=class_names,
    )
    plt.title("Pārpratumu matrica (Confusion Matrix)")
    plt.ylabel("Patiesā klase")
    plt.xlabel("Prognozētā klase")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()

    buffer = io.BytesIO()
    plt.savefig(buffer, format="png", dpi=150)
    plt.close()
    buffer.seek(0)
    return buffer


# Formatē šūnu ar kopīgiem stila parametriem
def apply_cell_style(cell, is_header: bool = False) -> None:
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if is_header:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    else:
        cell.font = ROW_FONT


# Izveido kopsavilkuma lapu ar galvenēm, ja tā vēl nepastāv
def ensure_summary_sheet(workbook: Workbook) -> None:
    if SUMMARY_SHEET_NAME not in workbook.sheetnames:
        sheet = workbook.create_sheet(SUMMARY_SHEET_NAME, 0)
        for col_index, header in enumerate(SUMMARY_HEADERS, start=1):
            cell = sheet.cell(row=1, column=col_index, value=header)
            apply_cell_style(cell, is_header=True)
            sheet.column_dimensions[get_column_letter(col_index)].width = 22
        sheet.row_dimensions[1].height = 20


# Pievieno vienu rezultātu rindu kopsavilkuma lapā
def append_summary_row(
    sheet,
    dataset_name: str,
    predict_duration: float,
    record_count: int,
    true_labels: np.ndarray,
    predicted_labels: np.ndarray,
    predicted_probabilities: np.ndarray,
    encoder: LabelEncoder,
) -> None:
    number_of_classes = len(encoder.classes_)
    auc_value = "N/A"
    try:
        if number_of_classes == 2:
            auc_value = round(roc_auc_score(true_labels, predicted_probabilities[:, 1]), 4)
        else:
            auc_value = round(
                roc_auc_score(true_labels, predicted_probabilities, multi_class="ovr", average="macro"), 4
            )
    except ValueError:
        pass

    row_values = [
        dataset_name,
        round(accuracy_score(true_labels, predicted_labels), 4),
        round(precision_score(true_labels, predicted_labels, average="macro", zero_division=0), 4),
        round(f1_score(true_labels, predicted_labels, average="macro"), 4),
        round(f1_score(true_labels, predicted_labels, average="weighted"), 4),
        auc_value,
        round(predict_duration, 4),
        record_count,
    ]

    next_row = sheet.max_row + 1
    for col_index, value in enumerate(row_values, start=1):
        cell = sheet.cell(row=next_row, column=col_index, value=value)
        apply_cell_style(cell)


# Izveido atsevišķu lapu datu kopai ar pilnu klasifikācijas pārskatu, feature importance un konfūzijas matricu
def write_detail_sheet(
    workbook: Workbook,
    dataset_name: str,
    predict_duration: float,
    record_count: int,
    true_labels: np.ndarray,
    predicted_labels: np.ndarray,
    predicted_probabilities: np.ndarray,
    encoder: LabelEncoder,
    feature_importance: pd.Series,
) -> None:
    # Noņem kopējo prefiksu, lai lapas nosaukumā paliktu tikai unikālā daļa
    common_prefix = "merged_no_nan_clean_headers_no_duplicates"
    display_name = dataset_name[len(common_prefix):].lstrip("_") if dataset_name.startswith(common_prefix) else dataset_name
    base_name = display_name[:31]
    sheet_name = base_name
    counter = 2
    while sheet_name in workbook.sheetnames:
        suffix = f"_{counter}"
        sheet_name = base_name[:31 - len(suffix)] + suffix
        counter += 1
    sheet = workbook.create_sheet(sheet_name)

    class_names = [str(c) for c in encoder.classes_]
    number_of_classes = len(class_names)
    report_dict = classification_report(
        true_labels, predicted_labels, target_names=class_names, zero_division=0, output_dict=True
    )

    auc_value = "N/A"
    try:
        if number_of_classes == 2:
            auc_value = f"{roc_auc_score(true_labels, predicted_probabilities[:, 1]):.4f}"
        else:
            auc_value = f"{roc_auc_score(true_labels, predicted_probabilities, multi_class='ovr', average='macro'):.4f}"
    except ValueError:
        pass

    summary_block = [
        ("Datukopas nosaukums", dataset_name),
        ("Ierakstu skaits", record_count),
        ("Prognozēšanas laiks (s)", f"{predict_duration:.4f}"),
        ("Precizitāte (Accuracy)", f"{accuracy_score(true_labels, predicted_labels):.4f}"),
        ("Precīzumspēja (Precision)", f"{precision_score(true_labels, predicted_labels, average='macro', zero_division=0):.4f}"),
        ("F1 rādītājs (F1 measure)", f"{f1_score(true_labels, predicted_labels, average='macro'):.4f}"),
        ("F1 Weighted", f"{f1_score(true_labels, predicted_labels, average='weighted'):.4f}"),
        ("AUC-ROC", auc_value),
    ]

    for row_index, (label, value) in enumerate(summary_block, start=1):
        label_cell = sheet.cell(row=row_index, column=1, value=label)
        value_cell = sheet.cell(row=row_index, column=2, value=value)
        apply_cell_style(label_cell, is_header=True)
        apply_cell_style(value_cell)

    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 22

    report_start_row = len(summary_block) + 2
    report_headers = ["Klase", "Precīzumspēja (precision)", "Pārklājums (recall)", "F1 rādītājs (F1 measure)", "Ierakstu skaits testa kopā"]

    for col_index, header in enumerate(report_headers, start=1):
        cell = sheet.cell(row=report_start_row, column=col_index, value=header)
        apply_cell_style(cell, is_header=True)
        sheet.column_dimensions[get_column_letter(col_index)].width = 22

    for row_offset, class_name in enumerate(class_names, start=1):
        metrics = report_dict.get(class_name, {})
        row_data = [
            class_name,
            f"{metrics.get('precision', 0):.4f}",
            f"{metrics.get('recall', 0):.4f}",
            f"{metrics.get('f1-score', 0):.4f}",
            int(metrics.get("support", 0)),
        ]
        for col_index, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=report_start_row + row_offset, column=col_index, value=value)
            apply_cell_style(cell)

    # Feature importance tabula — parāda pazīmju nozīmību konkrētajā datu kopā
    importance_start_row = report_start_row + len(class_names) + 2
    importance_headers = ["Pazīme", "Svarīgums (Svars)"]

    for col_index, header in enumerate(importance_headers, start=1):
        cell = sheet.cell(row=importance_start_row, column=col_index, value=header)
        apply_cell_style(cell, is_header=True)

    top_features = feature_importance.nlargest(10)
    for row_offset, (feature_name, importance_value) in enumerate(top_features.items(), start=1):
        name_cell = sheet.cell(row=importance_start_row + row_offset, column=1, value=feature_name)
        value_cell = sheet.cell(row=importance_start_row + row_offset, column=2, value=round(float(importance_value), 4))
        apply_cell_style(name_cell)
        apply_cell_style(value_cell)

    matrix_image_row = importance_start_row + 10 + 2
    image_buffer = build_confusion_matrix_image(true_labels, predicted_labels, class_names)
    excel_image = XlImage(image_buffer)
    excel_image.anchor = f"A{matrix_image_row}"
    sheet.add_image(excel_image)


# Saglabā pilnu novērtējumu Excel failā — kopsavilkuma rinda un detalizēta lapa
def save_results_to_excel(
    dataset_name: str,
    predict_duration: float,
    record_count: int,
    true_labels: np.ndarray,
    predicted_labels: np.ndarray,
    predicted_probabilities: np.ndarray,
    encoder: LabelEncoder,
    feature_importance: pd.Series,
) -> None:
    if RESULTS_FILE.exists():
        workbook = load_workbook(RESULTS_FILE)
    else:
        workbook = Workbook()
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

    ensure_summary_sheet(workbook)
    summary_sheet = workbook[SUMMARY_SHEET_NAME]

    append_summary_row(
        summary_sheet, dataset_name, predict_duration, record_count,
        true_labels, predicted_labels, predicted_probabilities, encoder,
    )
    write_detail_sheet(
        workbook, dataset_name, predict_duration, record_count,
        true_labels, predicted_labels, predicted_probabilities, encoder,
        feature_importance,
    )

    workbook.save(RESULTS_FILE)
    print(f"Rezultati saglabati: {RESULTS_FILE}")


# Drukā novērtēšanas metriku kopsavilkumu konsolē
def print_evaluation_metrics(
    true_labels: np.ndarray,
    predicted_labels: np.ndarray,
    predicted_probabilities: np.ndarray,
    encoder: LabelEncoder,
) -> None:
    class_names = [str(c) for c in encoder.classes_]
    number_of_classes = len(class_names)

    print(f"Accuracy : {accuracy_score(true_labels, predicted_labels):.4f}")
    print(f"Precision: {precision_score(true_labels, predicted_labels, average='macro', zero_division=0):.4f}")
    print(f"F1 macro : {f1_score(true_labels, predicted_labels, average='macro'):.4f}")
    print(f"F1 weight: {f1_score(true_labels, predicted_labels, average='weighted'):.4f}")

    try:
        if number_of_classes == 2:
            auc = roc_auc_score(true_labels, predicted_probabilities[:, 1])
        else:
            auc = roc_auc_score(
                true_labels, predicted_probabilities, multi_class="ovr", average="macro"
            )
        print(f"AUC-ROC  : {auc:.4f}")
    except ValueError:
        pass

    print("\n" + classification_report(
        true_labels, predicted_labels, target_names=class_names, zero_division=0,
    ))


# Drukā modeļa svarīgāko pazīmju rangu, lai novērtētu telemetrijas datu nozīmību
def print_feature_importance(model: XGBClassifier, feature_names: pd.Index) -> None:
    importance_series = pd.Series(model.feature_importances_, index=feature_names)
    top_features = importance_series.nlargest(10)
    print("\nTop-10 feature importance:")
    for feature_name, importance_value in top_features.items():
        print(f"  {feature_name:<40} {importance_value:.4f}")


# Apmāca jaunu XGBoost modeli uz visiem pieejamajiem datiem un saglabā to
def train_model(csv_path: str, target_column: str) -> None:
    stem = Path(csv_path).stem
    model_path = Path(f"{stem}.pkl")
    encoder_path = Path(f"{stem}_encoder.pkl")

    print(f"Ielādē: {Path(csv_path).name}")
    features, labels = load_dataset(csv_path, target_column)
    print(f"Rindas: {len(features):,}  |  Pazīmes: {features.shape[1]}  |  Klases: {labels.nunique()}\n")

    encoded_labels, encoder = encode_labels(labels, encoder_path)
    sample_weights = compute_sample_weights(encoded_labels)

    total_trees = XGBOOST_PARAMS["n_estimators"]
    model = XGBClassifier(**XGBOOST_PARAMS, callbacks=[TqdmProgressCallback(total_trees)])

    training_start = time.time()
    model.fit(features, encoded_labels, sample_weight=sample_weights, verbose=False)
    training_duration = time.time() - training_start

    # Noņem callback pirms saglabāšanas, lai izvairītos no serializācijas kļūdām
    model.set_params(callbacks=None)
    joblib.dump(model, model_path)

    print(f"\nApmācības laiks : {training_duration:.1f}s")
    print(f"Modelis saglabāts  : {model_path}")
    print(f"Enkoderis saglabāts: {encoder_path}")
    print_feature_importance(model, features.columns)


# Ielādē saglabāto modeli un novērtē to uz norādītās datu kopas
def evaluate_saved_model(csv_path: str, target_column: str, model_path: str) -> None:
    encoder_path = resolve_encoder_path(model_path)
    model = joblib.load(model_path)
    encoder = joblib.load(encoder_path)

    features, labels = load_dataset(csv_path, target_column)
    encoded_labels = encoder.transform(labels)

    predict_start = time.time()
    predicted_labels = model.predict(features)
    predicted_probabilities = model.predict_proba(features)
    predict_duration = time.time() - predict_start

    feature_importance = pd.Series(model.feature_importances_, index=features.columns)

    dataset_name = Path(model_path).stem
    print(f"Prognozēšanas laiks: {predict_duration:.4f}s / {len(features):,} ieraksti\n")
    print_evaluation_metrics(encoded_labels, predicted_labels, predicted_probabilities, encoder)
    save_results_to_excel(
        dataset_name, predict_duration, len(features),
        encoded_labels, predicted_labels, predicted_probabilities, encoder,
        feature_importance,
    )


# Ielādē saglabāto modeli un ģenerē prognozes CSV failam bez mērķa kolonnas
def predict_with_saved_model(csv_path: str, model_path: str) -> None:
    encoder_path = resolve_encoder_path(model_path)
    model = joblib.load(model_path)
    encoder = joblib.load(encoder_path)

    features = pd.read_csv(csv_path, low_memory=False)
    predicted_encoded = model.predict(features)
    predicted_labels = encoder.inverse_transform(predicted_encoded)

    output_dataframe = features.copy()
    output_dataframe["Predicted_Label"] = predicted_labels
    output_dataframe.to_csv("predictions.csv", index=False)

    print(pd.Series(predicted_labels).value_counts().to_string())


# Nolasa faila ceļu no lietotāja un pārbauda vai fails eksistē
def prompt_existing_file(prompt_text: str) -> str:
    while True:
        path = input(prompt_text).strip().strip('"')
        if Path(path).is_file():
            return path
        print(f"  Fails nav atrasts: {path}")


# Parāda interaktīvu izvēlni un izpilda lietotāja izvēlēto darbību
def run_interactive_menu() -> None:
    print("\n=== XGBoost IDS Klasifikators ===")
    print("  1. Apmācīt jaunu modeli")
    print("  2. Novērtēt saglabāto modeli")
    print("  3. Prognozēt (bez Label kolonnas)")
    print("  0. Iziet")
    print()

    choice = input("Izvēle: ").strip()

    if choice == "1":
        csv_path = prompt_existing_file("Apmācības CSV fails: ")
        target_column = input("Mērķa kolonna (noklusējums: Label): ").strip() or "Label"
        train_model(csv_path, target_column)

    elif choice == "2":
        csv_path = prompt_existing_file("Testa CSV fails: ")
        model_path = prompt_existing_file("Modeļa fails (.pkl): ")
        target_column = input("Mērķa kolonna (noklusējums: Label): ").strip() or "Label"
        evaluate_saved_model(csv_path, target_column, model_path)

    elif choice == "3":
        csv_path = prompt_existing_file("CSV fails (bez Label kolonnas): ")
        model_path = prompt_existing_file("Modeļa fails (.pkl): ")
        predict_with_saved_model(csv_path, model_path)

    elif choice == "0":
        return

    else:
        print("Nepareiza izvēle.")


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", choices=["train", "evaluate", "predict"], default=None)
    parser.add_argument("--csv", default=None)
    parser.add_argument("--target", default="Label")
    parser.add_argument("--model", default=None, help="Ceļš uz .pkl modeļa failu")
    return parser.parse_args()


if __name__ == "__main__":
    arguments = parse_arguments()

    # Ja nav padoti argumenti, palaiž interaktīvo izvēlni
    if arguments.mode is None:
        run_interactive_menu()
    elif arguments.mode == "train":
        train_model(arguments.csv, arguments.target)
    elif arguments.mode == "evaluate":
        evaluate_saved_model(arguments.csv, arguments.target, arguments.model)
    elif arguments.mode == "predict":
        predict_with_saved_model(arguments.csv, arguments.model)